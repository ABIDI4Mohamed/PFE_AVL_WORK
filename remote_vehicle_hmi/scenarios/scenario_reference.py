from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, List, Optional, Any

from openpyxl import load_workbook


class ScenarioReference:
    """
    Load scenario references from either:
    - CSV  (.csv)
    - Excel workbook (.xlsx / .xlsm)

    Supported schemas:

    1) Legacy fixed-target schema:
        scenario_id, step_id, time_s, steer_ref_deg, throttle_ref, brake_ref,
        speed_ref_mps, tol_steer_deg, tol_throttle, tol_brake, tol_speed_mps, phase

    2) V2 corridor-based schema:
        scenario_id, step_id, time_s, phase, steer_ref_deg, throttle_ref, brake_ref,
        speed_min_mps, speed_max_mps, tol_steer_deg, tol_throttle, tol_brake,
        eval_method

    Excel support:
    - One sheet per scenario is recommended (sheet title used as scenario_id if the
      sheet does not contain a scenario_id column).
    - The workbook may also include helper sheets like README; unsupported sheets
      are skipped automatically.
    - Optional mass-aware columns:
        mass_label, vehicle_mass_kg
    """

    LEGACY_REQUIRED_COLUMNS = [
        "scenario_id",
        "step_id",
        "time_s",
        "steer_ref_deg",
        "throttle_ref",
        "brake_ref",
        "speed_ref_mps",
        "tol_steer_deg",
        "tol_throttle",
        "tol_brake",
        "tol_speed_mps",
        "phase",
    ]

    V2_REQUIRED_COLUMNS = [
        "scenario_id",
        "step_id",
        "time_s",
        "phase",
        "steer_ref_deg",
        "throttle_ref",
        "brake_ref",
        "speed_min_mps",
        "speed_max_mps",
        "tol_steer_deg",
        "tol_throttle",
        "tol_brake",
    ]

    NUMERIC_COLUMNS = [
        "time_s",
        "steer_ref_deg",
        "throttle_ref",
        "brake_ref",
        "speed_ref_mps",
        "speed_min_mps",
        "speed_max_mps",
        "tol_steer_deg",
        "tol_throttle",
        "tol_brake",
        "tol_speed_mps",
        "vehicle_mass_kg",
    ]

    OPTIONAL_FILTER_COLUMNS = ["mass_label", "vehicle_mass_kg"]

    def __init__(self, source_path: str | Path) -> None:
        self.source_path = Path(source_path)
        self.rows: List[Dict[str, Any]] = []
        self.rows_by_scenario: Dict[str, List[Dict[str, Any]]] = {}
        self.schema: str = "unknown"
        self.format: str = "unknown"
        self._load()

    def _load(self) -> None:
        if not self.source_path.exists():
            raise FileNotFoundError(f"Scenario source not found: {self.source_path}")

        suffix = self.source_path.suffix.lower()
        if suffix == ".csv":
            self.format = "csv"
            self._load_csv()
        elif suffix in {".xlsx", ".xlsm"}:
            self.format = "excel"
            self._load_excel()
        else:
            raise ValueError(
                f"Unsupported scenario source format: {self.source_path.suffix}. "
                "Supported: .csv, .xlsx, .xlsm"
            )

        self.rows.sort(key=lambda r: (
            str(r.get("scenario_id", "")),
            self._safe_float(r.get("time_s", 0.0)),
        ))

        grouped: Dict[str, List[Dict[str, Any]]] = {}
        for row in self.rows:
            grouped.setdefault(str(row["scenario_id"]), []).append(row)
        self.rows_by_scenario = grouped

    def _load_csv(self) -> None:
        with self.source_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)

            if reader.fieldnames is None:
                raise ValueError("Scenario CSV is empty or has no header.")

            fields = self._normalize_header(reader.fieldnames)
            schema = self._detect_schema(fields, allow_inferred_scenario=False)
            self.schema = schema

            for line_idx, row in enumerate(reader, start=2):
                cleaned = self._clean_row(
                    row=dict(row),
                    schema=schema,
                    context=f"CSV line {line_idx}",
                    default_scenario_id=None,
                )
                if cleaned is not None:
                    self.rows.append(cleaned)

    def _load_excel(self) -> None:
        wb = load_workbook(self.source_path, data_only=True, read_only=True)
        schemas_seen = set()

        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)

            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue

            if header_row is None:
                continue

            raw_header = ["" if v is None else str(v).strip() for v in header_row]
            if not any(raw_header):
                continue

            fields = self._normalize_header(raw_header)

            try:
                schema = self._detect_schema(fields, allow_inferred_scenario=True)
            except ValueError:
                # Skip helper sheets like README / notes / calibration aids
                continue

            schemas_seen.add(schema)

            for line_idx, values in enumerate(rows_iter, start=2):
                if values is None:
                    continue

                row_dict = {}
                for i, field in enumerate(fields):
                    if not field:
                        continue
                    value = values[i] if i < len(values) else None
                    row_dict[field] = value

                if not any(v not in (None, "") for v in row_dict.values()):
                    continue

                cleaned = self._clean_row(
                    row=row_dict,
                    schema=schema,
                    context=f"Sheet '{ws.title}' line {line_idx}",
                    default_scenario_id=ws.title,
                )
                if cleaned is not None:
                    self.rows.append(cleaned)

        if not self.rows:
            raise ValueError(
                "No valid scenario sheet/table found in workbook. "
                "Make sure each scenario sheet has a table header matching the expected schema."
            )

        if len(schemas_seen) == 1:
            self.schema = next(iter(schemas_seen))
        else:
            self.schema = "mixed"

    def _normalize_header(self, fields: List[Any]) -> List[str]:
        normalized: List[str] = []
        for x in fields:
            s = "" if x is None else str(x).strip()
            normalized.append(s)
        return normalized

    def _detect_schema(self, fields: List[str], allow_inferred_scenario: bool) -> str:
        effective_fields = set(fields)
        if allow_inferred_scenario and "scenario_id" not in effective_fields:
            effective_fields.add("scenario_id")

        legacy_missing = [c for c in self.LEGACY_REQUIRED_COLUMNS if c not in effective_fields]
        v2_missing = [c for c in self.V2_REQUIRED_COLUMNS if c not in effective_fields]

        if not legacy_missing:
            return "legacy"
        if not v2_missing:
            return "v2"

        raise ValueError(
            "Scenario source does not match a supported schema.\n"
            f"Legacy missing: {legacy_missing}\n"
            f"V2 missing: {v2_missing}"
        )

    def _clean_row(
        self,
        row: Dict[str, Any],
        schema: str,
        context: str,
        default_scenario_id: Optional[str],
    ) -> Optional[Dict[str, Any]]:
        cleaned = dict(row)

        # Infer scenario_id from sheet name for Excel sheets if absent/blank
        scenario_id = cleaned.get("scenario_id")
        if scenario_id in (None, ""):
            if default_scenario_id is not None:
                cleaned["scenario_id"] = default_scenario_id
            else:
                raise ValueError(f"Missing scenario_id at {context}")

        # Required string-ish fields
        cleaned["scenario_id"] = str(cleaned["scenario_id"]).strip()
        cleaned["step_id"] = str(cleaned.get("step_id", "")).strip()
        cleaned["phase"] = str(cleaned.get("phase", "")).strip()

        # Optional mass fields
        if "mass_label" in cleaned and cleaned["mass_label"] not in (None, ""):
            cleaned["mass_label"] = str(cleaned["mass_label"]).strip()
        if "vehicle_mass_kg" in cleaned and cleaned["vehicle_mass_kg"] in ("", None):
            cleaned["vehicle_mass_kg"] = None

        # Convert numeric fields only if present and not blank
        for key in self.NUMERIC_COLUMNS:
            if key in cleaned and cleaned[key] not in (None, ""):
                try:
                    cleaned[key] = float(cleaned[key])
                except (TypeError, ValueError):
                    raise ValueError(
                        f"Invalid numeric value for '{key}' at {context}: {cleaned.get(key)!r}"
                    )

        return cleaned

    def _safe_float(self, value: Any) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def list_scenarios(self) -> List[str]:
        return sorted(self.rows_by_scenario.keys())

    def has_scenario(self, scenario_id: str) -> bool:
        return scenario_id in self.rows_by_scenario

    def _filter_rows(
        self,
        rows: List[Dict[str, Any]],
        mass_label: Optional[str] = None,
        vehicle_mass_kg: Optional[float] = None,
    ) -> List[Dict[str, Any]]:
        filtered = rows

        if mass_label is not None:
            filtered = [
                r for r in filtered
                if ("mass_label" not in r) or str(r.get("mass_label", "")).strip() == str(mass_label).strip()
            ]

        if vehicle_mass_kg is not None:
            target = float(vehicle_mass_kg)
            filtered = [
                r for r in filtered
                if ("vehicle_mass_kg" not in r) or (
                    r.get("vehicle_mass_kg") is not None and float(r["vehicle_mass_kg"]) == target
                )
            ]

        return filtered

    def get_rows_for_scenario(
        self,
        scenario_id: str,
        mass_label: Optional[str] = None,
        vehicle_mass_kg: Optional[float] = None,
    ) -> List[Dict[str, Any]]:
        if scenario_id not in self.rows_by_scenario:
            raise KeyError(f"Unknown scenario_id: {scenario_id}")

        rows = self.rows_by_scenario[scenario_id]
        rows = self._filter_rows(rows, mass_label=mass_label, vehicle_mass_kg=vehicle_mass_kg)
        return rows

    def get_reference_at_time(
        self,
        scenario_id: str,
        t_s: float,
        mass_label: Optional[str] = None,
        vehicle_mass_kg: Optional[float] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        Zero-order hold:
        returns the last row whose time_s <= t_s.
        If t_s is before the first row, returns the first row.
        """
        rows = self.get_rows_for_scenario(
            scenario_id,
            mass_label=mass_label,
            vehicle_mass_kg=vehicle_mass_kg,
        )
        if not rows:
            return None

        if t_s <= float(rows[0]["time_s"]):
            return rows[0]

        current = rows[0]
        for row in rows:
            if float(row["time_s"]) <= t_s:
                current = row
            else:
                break
        return current

    def get_next_reference_after(
        self,
        scenario_id: str,
        t_s: float,
        mass_label: Optional[str] = None,
        vehicle_mass_kg: Optional[float] = None,
    ) -> Optional[Dict[str, Any]]:
        rows = self.get_rows_for_scenario(
            scenario_id,
            mass_label=mass_label,
            vehicle_mass_kg=vehicle_mass_kg,
        )
        for row in rows:
            if float(row["time_s"]) > t_s:
                return row
        return None

    def get_phase_at_time(
        self,
        scenario_id: str,
        t_s: float,
        mass_label: Optional[str] = None,
        vehicle_mass_kg: Optional[float] = None,
    ) -> Optional[str]:
        ref = self.get_reference_at_time(
            scenario_id,
            t_s,
            mass_label=mass_label,
            vehicle_mass_kg=vehicle_mass_kg,
        )
        if ref is None:
            return None
        return str(ref["phase"])

    def get_duration(
        self,
        scenario_id: str,
        mass_label: Optional[str] = None,
        vehicle_mass_kg: Optional[float] = None,
    ) -> float:
        rows = self.get_rows_for_scenario(
            scenario_id,
            mass_label=mass_label,
            vehicle_mass_kg=vehicle_mass_kg,
        )
        if not rows:
            return 0.0
        return float(rows[-1]["time_s"])

    def as_dict(self) -> Dict[str, List[Dict[str, Any]]]:
        return self.rows_by_scenario
